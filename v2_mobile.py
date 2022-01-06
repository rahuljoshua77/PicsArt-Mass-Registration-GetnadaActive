import pandas
from openpyxl import load_workbook
import re
import requests,random,json,os
import pandas as pd


cwd = os.getcwd()
from bs4 import BeautifulSoup
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from selenium.webdriver.chrome.service import Service
from multiprocessing import Pool
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.support import expected_conditions as EC

name_card = random.choice(["Bellamy","Nurman","Herman","Michael","Michelle","Jeniffer","Robby","Eva","Nurlis","Yudi","Franky","Rudi","Kurdiawan"])

mobile_emulation = {
    "deviceMetrics": { "width": 360, "height": 650, "pixelRatio": 3.4 },
    }
brave_path = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"
driver_path= f"{cwd}\\chromedriver.exe"
firefox_options = webdriver.ChromeOptions()
firefox_options.add_argument('--no-sandbox')

firefox_options.headless = False
firefox_options.add_argument('--disable-setuid-sandbox')
firefox_options.add_argument('disable-infobars')
firefox_options.add_argument('--ignore-certifcate-errors')
firefox_options.add_argument('--ignore-certifcate-errors-spki-list')
firefox_options.add_argument("--incognito")
firefox_options.add_argument('--no-first-run')
firefox_options.add_argument('--disable-dev-shm-usage')
firefox_options.add_argument("--disable-infobars")
firefox_options.add_argument("--disable-extensions")
firefox_options.add_argument("--disable-popup-blocking")
firefox_options.add_argument('--log-level=3')
firefox_options.add_argument("--window-size=500,1090")
firefox_options.add_argument('--disable-blink-features=AutomationControlled')
firefox_options.add_experimental_option("useAutomationExtension", False)
firefox_options.add_experimental_option("excludeSwitches",["enable-automation"])
firefox_options.add_experimental_option('excludeSwitches', ['enable-logging'])
firefox_options.add_argument('--disable-notifications')
from selenium.webdriver.common.action_chains import ActionChains
firefox_options.binary_location = brave_path
random_angka = random.randint(100,999)
random_angka_dua = random.randint(10,99)
header = {"accept-encoding": "gzip, deflate",
         "content-type": "application/json; charset=utf-8",
    "accept-language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
    "content-length": "76",
    "content-type": "application/json",
    "cookie": 'top-navigation-experiment-variant=0; forcedRegABExpId=HMAxRyAXROWAsfH9HxDqVQ; forcedRegABVariant=0; _gcl_au=1.1.1528467627.1640452192; _ga=GA1.2.1177011315.1640452192; _gid=GA1.2.971393706.1640452192; _fbp=fb.1.1640452193216.1940445925; paa-did=a.c.kxm2vnuy.66b8711a-9ea5-4d57-bb11-ce8591ea22b9; _pin_unauth=dWlkPU9EYzRabVl3TVRndE5UVmlNeTAwWkRJNUxXRTNNREl0TlRZd1l6ZGhNR1UyTW1ZMw; AF_SYNC=1640452195103; afUserId=4322f47b-944a-4949-bc48-fad52478a2b5-p; g_state={"i_p":1640459400899,"i_l":1}; OptanonAlertBoxClosed=2021-12-25T17:10:04.487Z; sid=s%3AEw0vmvGzZbOZkonmA9msuUBGDxkHm71q.zeaoZPlrm06pe4kLdxKcsJK8I6fiRQKL9oaa5BxNt9A; isOpened=true; IR_gbd=picsart.com; _fbc=fb.1.1640518552375.IwAR3fyXlJ1PSgh3jKvyUw3jP6dPZ7VRqK6v-By_hbnaO38ycJKzb8RixQifs; ab.storage.deviceId.4fc46d16-14ad-4944-ba4b-c874f391cb00=%7B%22g%22%3A%22cefd7c95-4389-8ecf-e619-58d6aff13614%22%2C%22c%22%3A1640452254332%2C%22l%22%3A1640518552478%7D; ab.storage.userId.4fc46d16-14ad-4944-ba4b-c874f391cb00=%7B%22g%22%3A%22378148284058101%22%2C%22c%22%3A1640452287372%2C%22l%22%3A1640518552479%7D; IR_PI=30f60a67-65a6-11ec-8fd8-2b13cb31605b%7C1640604967147; __zlcmid=17jkaDwghOocvF9; currentLanguage=en; tatari-cookie-test=11201914; tatari-user-cookie=378148284058101; IR_11703=1640518746737%7C0%7C1640518746737%7C%7C; OptanonConsent=isGpcEnabled=0&datestamp=Sun+Dec+26+2021+18%3A39%3A07+GMT%2B0700+(Waktu+Indonesia+Barat)&version=6.20.0&isIABGlobal=false&hosts=&consentId=cb7197ac-5209-4679-b70b-8a58e47205f8&interactionCount=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0002%3A1%2CC0003%3A1%2CC0004%3A1&geolocation=%3B&AwaitingReconsent=false; tatari-session-cookie=538817ee-e7fc-f812-8174-85987aa7c54a; ab.storage.sessionId.4fc46d16-14ad-4944-ba4b-c874f391cb00=%7B%22g%22%3A%22b64d4c64-d3c8-ba89-963d-49e761f9a048%22%2C%22e%22%3A1640520601808%2C%22c%22%3A1640518552475%2C%22l%22%3A1640518801808%7D',
    "deviceid": 'a.c.kxm2vnuy.66b8711a-9ea5-4d57-bb11-ce8591ea22b9',
    "origin": "https://picsart.com",
    "referer": "https://picsart.com/universe-trial?fbclid=IwAR3fyXlJ1PSgh3jKvyUw3jP6dPZ7VRqK6v-By_hbnaO38ycJKzb8RixQifs",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1'
    }
def xpath_el(el):
    element_all = wait(browser,15).until(EC.presence_of_element_located((By.XPATH, el)))
    browser.execute_script("arguments[0].scrollIntoView();", element_all)
    return browser.execute_script("arguments[0].click();", element_all)

def xpath_ex(el):
    element_all = wait(browser,0.3).until(EC.presence_of_element_located((By.XPATH, el)))
    browser.execute_script("arguments[0].scrollIntoView();", element_all)
    return browser.execute_script("arguments[0].click();", element_all)

def sign_up(email, password,new_cc):
    global browser
    get_cc = new_cc.replace("Live | ","").replace(" [BIN: - - - ][GATE:01] @/ChkNET-ID","")
    get_data = get_cc.split("|")
    number_card = get_data[0]
    month = get_data[1]
    year = get_data[2]
    get_year = str(year).split("20")
    expired_card = str(month)+f"{get_year[1]}"
    security_card = get_data[3]

    #firefox_options.add_experimental_option("mobileEmulation", mobile_emulation)
    firefox_options.add_argument(f"user-agent=Mozilla/5.0 (iPad; CPU OS 11_0 like Mac OS X) AppleWebKit/604.1.34 (KHTML, like Gecko) Version/11.0 Mobile/15A5341f Safari/604.1")
    browser = webdriver.Chrome(options=firefox_options,executable_path=driver_path)
    browser.execute_script("document.body.style.zoom='zoom 90%'")
    browser.get('https://picsart.com/universe-trial')
    print("[*] Trying to Creating New Account...!")
    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Trying to Creating New Account...! \n')
     
    element_all = wait(browser,0.3).until(EC.presence_of_element_located((By.XPATH,'(//span[contains(@class,"button-text")])[1]')))
    
    browser.execute_script("arguments[0].click();", element_all)
    try:
        xpath_el('//button[@id="onetrust-accept-btn-handler"]')
        #xpath_el('//a[contains(@class,"switchAction")]')
    except:
        pass
    try:
        email_input = wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//input[@name="email"]')))
        email_input.send_keys(email)
        sleep(0.5)
        pw_input = wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//input[@name="password"]')))
        pw_input.send_keys(password)
        wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//button[@type="submit"]'))).click()
        n = 1
        while True:
            if n == 3:
                break
            try:
                wait(browser,0.5).until(EC.presence_of_element_located((By.XPATH, f'//input[@aria-label="Pay using credit card"]'))).click()

                break
            except:
                try:
                    wait(browser,0.5).until(EC.presence_of_element_located((By.XPATH, f'//button[@type="submit"]'))).click()
                except:
                    try:
                        wait(browser,0.5).until(EC.presence_of_element_located((By.XPATH, f"//span[contains(text(),'Something went wrong!')]")))
                        print("[*] Run Again! Change IP!")
                        break
                    except:
                        pass
        while True:
                
            if n == 5:
                print("[*] Verification Failed!")
                with open('log.txt','a',encoding='utf-8') as f: f.write(f'Verification Failed!\n')
                break
            URL = f'https://getnada.com/api/v1/inboxes/{email}'
            r = requests.get(URL).json()
            #getting the latest message
            
            try:
                global uid
                uid = r['msgs'][0]['uid']
            
                mes = requests.get(f'https://getnada.com/api/v1/messages/html/{uid}')
                mes1 = BeautifulSoup(mes.content,'html.parser')
                sleep(1)
                get_data = mes1.prettify()
                get_data = get_data.split('href="https://picsart.com/activate/')
                get_data = get_data[1].split('style="display: block; padding-left: 45px;')
                get_data = get_data[0]
                get_data = get_data.split('"')
                url_activation = f'https://picsart.com/activate/{get_data[0]}'
                print(f'[*] Creating New Account Success!')
                with open('log.txt','a',encoding='utf-8') as f: f.write(f'Creating New Account Success! \n')
                requests.get(f'https://picsart.com/activate/{get_data[0]}')
                print("[*] Verification Email Success!")
                with open('log.txt','a',encoding='utf-8') as f: f.write(f'Verification Email Success!\n')
                print("[*] Please wait while you are now trying to Execution the PicsArt Account with this Credit Card details...")
                with open('log.txt','a',encoding='utf-8') as f: f.write(f'Please wait while you are now trying to Execution the PicsArt Account with this Credit Card details... \n')
                #auto_pay(email, password)
                break
            except IndexError:
                print("[*] Your Email doesn't have a new message, Reload!")
                with open('log.txt','a',encoding='utf-8') as f: f.write("Your Email doesn't have a new message, Reload! \n")
                n = n+1
                sleep(2)
        print(f"[*] {number_card}|{month}|{year}|{security_card}|{name_card}|")
        wait(browser, 40).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"(//iframe[contains(@src,'https://checkoutshopper-live.adyen.com/')])[1]")))
        input_number_card = wait(browser,30).until(EC.presence_of_element_located((By.XPATH, f'//input[@id="encryptedCardNumber"]')))
        input_number_card.send_keys(number_card)
        sleep(0.5)
        browser.switch_to.default_content()
        #wait(browser, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"(//iframe[contains(@src,'https://checkoutshopper-live.adyen.com/')])")))
        wait(browser, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"(//iframe[contains(@src,'https://checkoutshopper-live.adyen.com/')])[2]")))
        input_expired_card = wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//input[@id="encryptedExpiryDate"]')))
        input_expired_card.send_keys(expired_card)
        sleep(0.5)
        browser.switch_to.default_content()
        #  wait(browser, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"(//iframe[contains(@src,'https://checkoutshopper-live.adyen.com/')])")))
        wait(browser, 10).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"(//iframe[contains(@src,'https://checkoutshopper-live.adyen.com/')])[3]")))
        input_security_card = wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//input[@id="encryptedSecurityCode"]')))
        input_security_card.send_keys(security_card)
        sleep(2)
        browser.switch_to.default_content()
        input_name_card = wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//input[@placeholder="Name on Card"]')))
        input_name_card.send_keys(name_card)
        sleep(0.5)
        n = 1
        try:
            wait(browser,10).until(EC.presence_of_element_located((By.XPATH, f'//button[@id="onetrust-accept-btn-handler"]'))).click()
        except:
            pass
        try:
            xpath_ex('//button[@data-test="checkout-form-payment-button-adyen"]')
            
            el = wait(browser,0.3).until(EC.presence_of_element_located((By.XPATH, f'//button[@data-test="checkout-form-payment-button-adyen"]')))
            el.click()
            action = ActionChains(browser)
            for i in range(0,3):               
                try:
                    action.double_click(on_element = el)
                except:
                    pass
        except:
            pass
        sleep(5)
        try:
            el = wait(browser,35).until(EC.presence_of_element_located((By.XPATH, f'//h3[@data-test="welcome-title"]'))).text
          
            while True:
                if n == 5:
            
                    print(f"[*] Unfortunately... the Auto Payment & Invoice was Failed :(")
                    print(f"[*] Your previous Credit Card was successfully used and deleted from cc.txt")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Your previous Credit Card was successfully used and deleted from cc.txt \n')
                    print(f"[*] And here are details of your previous Credit Card")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'And here are details of your previous Credit Card \n')
                    print(f"[*] {number_card}|{month}|{year}|{security_card}|{name_card}|")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'{number_card}|{month}|{year}|{security_card}|{name_card}| \n')
                    print(f"[*] Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User\n')
                    print(f"[*] You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx\n')
                  
                    lst = [f"{email}|{password}"]
                    book = load_workbook('resultfailed.xlsx')
                    writer = pandas.ExcelWriter('resultfailed.xlsx', engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    df = pd.DataFrame(lst)
                    for sheetname in writer.sheets:
                        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

                    writer.save()
                    with open('resultfailed.txt','a') as f:
                        f.write(f"{number}. {email}|{password}")
                    dats.remove(new_cc)
                    with open('cc.txt','w',encoding='utf-8') as f: f.write(f'') 
                    for m in dats[:]: 
                        with open('cc.txt','a',encoding='utf-8') as f: f.write(f'{m}\n')
                    sleep(3)
                    break
                URL = f'https://getnada.com/api/v1/inboxes/{email}'
                r = requests.get(URL).json()
                
                try:
                
                    uid = r['msgs'][0]['uid']
                
                    mes = requests.get(f'https://getnada.com/api/v1/messages/html/{uid}')
                    mes1 = BeautifulSoup(mes.content,'html.parser')
                    get_data = mes1.prettify()
                    
                    data = get_data.split('''<p style="margin:44px 0 16px 0;color:#080808;font-size:24px;line-height:32px;font-weight:600;font-family:'Open Sans',Arial,sans-serif;">''')
                    data = data[1].split("Picsart Gold!")
                    #print(f"[*] {data[0].strip()} Picsart Gold!")
                    data = get_data.split('Your Picsart Gold subscription')
                    data = data[1].split("To avoid being charged, you must")
                    #print(f"[*] Your Picsart Gold subscription {data[0].strip()}")
                    date = data[0].split('on')
                    userdata = get_data.split('Username:')
                    userdata = userdata[1].split('''<td style="width:50%;vertical-align:top;padding:16px 0 0 16px;font-size:14px;line-height:20px;color:#080808;font-family:'Open Sans',Arial,sans-serif;">''')
                    userdata = userdata[1].split("</td>")
                    data = get_data.split('Date of purchase:')
                    datas = data[1].split(''' <td style="width:50%;vertical-align:top;padding:16px 0 0 16px;font-size:14px;line-height:20px;color:#080808;font-family:'Open Sans',Arial,sans-serif;">''')
                    data = datas[1].split('</tr>')
                    data = data[0].split("</td>")
                    date_buy = data[0].strip()
                    date_end = date[1].strip()
                    username = userdata[0].strip()
                    print("[*] Welcome To Gold!")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Welcome To Gold! \n')
                    print("[*] Auto Payment Success!")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Auto Payment Success! \n')
                    print(f"[*] Invoice Success!")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Invoice Success! \n')
                    print(f"[*] Your previous Credit Card was successfully used and deleted from cc.txt")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Your previous Credit Card was successfully used and deleted from cc.txt \n')
                    print(f"[*] And here are details of your previous Credit Card")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'And here are details of your previous Credit Card \n')
                    print(f"[*] {number_card}|{month}|{year}|{security_card}|{name_card}|")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'{number_card}|{month}|{year}|{security_card}|{name_card}| \n')
                    print(f"[*] And these are few details of your PicsArt Account")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'And these are few details of your PicsArt Account \n')
                    date_end = date_end.replace('.','')
                    data_month_buy = date_buy.split(",")
                    data_month_buy = data_month_buy[0].split(" ")
                    data_month_buy = data_month_buy[0]
                    data_month_end = date_end.split(",")
                    data_month_end = data_month_end[0].split(" ")
                    data_month_end = data_month_end[0]
                    list_month = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
                    get_idx_month_buy = list_month.index(data_month_buy)+1
                    get_month_int_buy = "0"+str(get_idx_month_buy)
                    get_idx_month_end= list_month.index(data_month_end)+1
                    get_month_int_end = "0"+str(get_idx_month_end)
                    date_buy_tgl = re.findall(r'\d+',date_buy)
                    date_end_tgl = re.findall(r'\d+',date_end)
                    date_buy_tgl = re.findall(r'\d+',date_buy)
                    date_end_tgl = re.findall(r'\d+',date_end)
                    print(f"[*] Username: {username} | Start: {date_buy_tgl[0]}-{get_month_int_buy}-2022 ({date_buy}) | Exp: {date_end_tgl[0]}-{get_month_int_end}-2022 ({date_end}) |")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Username: {username} | Start: {date_buy_tgl[0]}-{get_month_int_buy}-2022 ({date_buy}) | Exp: {date_end_tgl[0]}-{get_month_int_end}-2022 ({date_end}) | \n')
                     
                    with open('resultsuccess.txt','a') as f:
                        f.write(f"{email}|{password}|{date_buy_tgl[0]}{get_month_int_buy}-2022 ({date_buy})|{date_end_tgl[0]}-{get_month_int_end}-2022 ({date_end})|{number_card}|{month}|{year}|{security_card}|{name_card}|Welcome To Gold!|\n")
                  
                    book = load_workbook('resultsuccess.xlsx')
                    writer = pandas.ExcelWriter('resultsuccess.xlsx', engine='openpyxl')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                 
                    lst = [f"{email}|{password}|{date_buy_tgl[0]}-{get_month_int_buy}-2022 ({date_buy})|{date_end_tgl[0]}-{get_month_int_end}-2022 ({date_end})|{number_card}|{month}|{year}|{security_card}|{name_card}|Welcome To Gold!|"]
                   
                    df = pd.DataFrame(lst)
                    for sheetname in writer.sheets:
                        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

                    writer.save()
                    dats.remove(new_cc)
                    with open('cc.txt','w',encoding='utf-8') as f: f.write(f'')
                    for m in dats[:]:
                        with open('cc.txt','a',encoding='utf-8') as f: f.write(f'{m}\n')

                    print(f"[*] Great Job! The data now was saved to resultsuccess.txt & resultsuccess.xlsx")
                    with open('log.txt','a',encoding='utf-8') as f: f.write(f'Great Job! The data now was saved to resultsuccess.txt & resultsuccess.xlsx \n')
                    break

                except IndexError:
                    #print("[*] Your Email doesn't have a new message, Reload!")
                    n = n+1

             
        except Exception as e:
          
            
            print(f"[*] Unfortunately... the Auto Payment & Invoice was Failed :(")
            print(f"[*] Your previous Credit Card was successfully used and deleted from cc.txt")
            with open('log.txt','a',encoding='utf-8') as f: f.write(f'Your previous Credit Card was successfully used and deleted from cc.txt \n')
            print(f"[*] And here are details of your previous Credit Card")
            with open('log.txt','a',encoding='utf-8') as f: f.write(f'And here are details of your previous Credit Card \n')
            print(f"[*] {number_card}|{month}|{year}|{security_card}|{name_card}|")
            with open('log.txt','a',encoding='utf-8') as f: f.write(f'{number_card}|{month}|{year}|{security_card}|{name_card}| \n')
            print(f"[*] Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User")
            with open('log.txt','a',encoding='utf-8') as f: f.write(f'Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User\n')
            print(f"[*] You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx")
            with open('log.txt','a',encoding='utf-8') as f: f.write(f'You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx\n')
            lst = [f"{email}|{password}"]
            book = load_workbook('resultfailed.xlsx')
            writer = pandas.ExcelWriter('resultfailed.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df = pd.DataFrame(lst)
            for sheetname in writer.sheets:
                df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

            writer.save()
            with open('resultfailed.txt','a') as f:
                f.write(f"{email}|{password}")
            dats.remove(new_cc)
            with open('cc.txt','w',encoding='utf-8') as f: f.write(f'') 
            for m in dats[:]: 
                with open('cc.txt','a',encoding='utf-8') as f: f.write(f'{m}\n')
            

    except Exception as e:
       
        print(f"[*] Unfortunately... the Auto Payment & Invoice was Failed :(")
        print(f"[*] Your previous Credit Card was successfully used and deleted from cc.txt")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'Your previous Credit Card was successfully used and deleted from cc.txt \n')
        print(f"[*] And here are details of your previous Credit Card")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'And here are details of your previous Credit Card \n')
        print(f"[*] {number_card}|{month}|{year}|{security_card}|{name_card}|")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'{number_card}|{month}|{year}|{security_card}|{name_card}| \n')
        print(f"[*] Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'Error Verifying reCAPTCHA| Your Payment was Failed| Retry Limitation User\n')
        print(f"[*] You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'You Can Try it Again Later! The data now was saved to resultfailed.txt & resultfailed.xlsx\n')
        lst = [f"{email}|{password}"]
        book = load_workbook('resultfailed.xlsx')
        writer = pandas.ExcelWriter('resultfailed.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df = pd.DataFrame(lst)
        for sheetname in writer.sheets:
            df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

        writer.save()
        with open('resultfailed.txt','a') as f:
            f.write(f"{number}. {email}|{password}")
        dats.remove(new_cc)
        with open('cc.txt','w',encoding='utf-8') as f: f.write(f'')
        for m in dats[:]:
            with open('cc.txt','a',encoding='utf-8') as f: f.write(f'{m}\n')
        sleep(3)

    #     browser.save_screenshot("errr_IP.png")

        #$### with open('log.txt','a',encoding='utf-8') as f: f.write(f' \n')

def main():
    global dats
    print("[*] Auto Verification & Auto Payment PicsArt Gold Premium 3 Months with collaboration Brave Browser x GetNada x NamsoGen x MrChecker")
    with open('log.txt','a',encoding='utf-8') as f: f.write(f'[*] Auto Verification & Auto Payment PicsArt Gold Premium 3 Months with collaboration Brave Browser x GetNada x NamsoGen x MrChecker\n')
    print("[*] URL Activation: https://picsart.com/universe-trial | BIN: 52297444xxxxxxxx")
    with open('log.txt','a',encoding='utf-8') as f: f.write(f'URL Activation: https://picsart.com/universe-trial | BIN: 52297444xxxxxxxx\n')
    password = "panda123"
    file_list_akun = "cc.txt"
    myfile_akun = open(f"{cwd}\\{file_list_akun}","r")
    akun = myfile_akun.read()
    dats = akun.split("\n")
    global number
    number = 1
    for i in dats:
        email = input("[*] Email (example: test123asa@getairmail.com): ")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'Email (example: test123asa@getairmail.com): {email}\n')
        print("[*] Your Default Password is: panda123")
        with open('log.txt','a',encoding='utf-8') as f: f.write(f'Your Default Password is: panda123\n')
        sign_up(email, password, i)
        #os\.system\("notepad\.exe cc\.txt"\)
        number = number + 1

main()
